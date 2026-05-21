import streamlit as st
import pandas as pd
from datetime import datetime
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Control Stoc real Arcaprod", page_icon="🏗️", layout="wide")

CENTRALIZATOR_FILE = "centralizator_arcaprod.xlsx"
CONFIG_STOCURI_FILE = "config_stocuri.xlsx"
EMAIL_DESTINATAR = "sageata14@gmail.com"

# Parola pentru panoul de Admin
PAROLA_ADMIN = "Arcaprod2026"

# Nomenclatoare
ANGAJATI = [
    "Andrei Barbuceanu", "Catalin", "George B", "Ionut R", "Cornel I", 
    "Marian D (Vampirul)", "Dumitru U (Mitel)", "Andrei D (Alifie)", 
    "Eugen T", "Ion T (Nelu)", "Marius V", "Gabriel V", "Iulian G", 
    "Geovani B", "Constantin U (Costi)", "Mircea R", "Aashish", 
    "Andrei B", "Ionut B", "Valentin"
]
TIP_MATERIAL = ["Profil Aluminiu", "Feronerie", "Accesorii", "Consumabile", "Sticlă"]
UNITATI_MASURA = ["cutii", "buc", "bax", "ml", "set", "bare"]

# Structură istoric conform cerințelor din fabrică
COLOANE_CENTRALIZATOR = ["Timestamp", "Angajat", "Tip Material", "Denumire Material", "Stoc Initial", "Cantitate Ceruta", "Stoc Ramas", "UM", "Status Stoc", "Observatii"]
COLOANE_STOCURI = ["Denumire Material", "Stoc Actual", "UM"]

# --- FUNCȚIE PENTRU TRIMITERE EMAIL ---
def trimite_email_cu_atansamente(fisiere):
    EMAIL_EXPEDIATOR = "adresa_ta_de_gmail@gmail.com"
    PAROLA_APLICATIE = "parola_ta_de_aplicatie_gmail" 
    
    if EMAIL_EXPEDIATOR == "adresa_ta_de_gmail@gmail.com":
        st.error("⚠️ Trebuie să configurezi adresa de email și parola de aplicație în cod pentru a trimite!")
        return False

    msg = MIMEMultipart()
    msg['From'] = EMAIL_EXPEDIATOR
    msg['To'] = EMAIL_DESTINATAR
    msg['Subject'] = f"📊 Raport Stocuri & Cereri Arcaprod - {datetime.now().strftime('%d-%m-%Y')}"
    
    corp_email = "Bună ziua,\n\nAtasat găsiți rapoartele actualizate privind stocurile și registrul de cereri din fabrică.\n\nO zi bună!"
    msg.attach(MIMEText(corp_email, 'plain'))
    
    for cale_fisier in fisiere:
        if os.path.exists(cale_fisier):
            nume_fisier = os.path.basename(cale_fisier)
            with open(cale_fisier, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename= {nume_fisier}")
                msg.attach(part)
                
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(EMAIL_EXPEDIATOR, PAROLA_APLICATIE)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        st.error(f"Eroare la trimiterea emailului: {e}")
        return False

# --- FUNCȚIE DE SALVARE EXCEL STILIZAT ---
def salveaza_excel_stilizat(df_date, cale_fisier, titlu_raport):
    wb = Workbook()
    ws = wb.active
    ws.title = "Date_Arcaprod"
    ws.views.sheetView[0].showGridLines = True
    
    ws["A1"] = titlu_raport.upper()
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F497D")
    
    ws["A2"] = f"Generat la data: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")
    
    headers = list(df_date.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    subtire = Side(border_style="thin", color="D9D9D9")
    contur_celula = Border(left=subtire, right=subtire, top=subtire, bottom=subtire)
    
    for row_num, row_data in enumerate(df_date.values, 5):
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = str(val) if headers[col_num-1] == "Timestamp" else val
            cell.font = Font(name="Arial", size=10)
            cell.border = contur_celula
            
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2, 3]: continue
            if cell.value: max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(cale_fisier)

# --- FUNCȚIE AJUTĂTOARE: CITEȘTE COLOANELE CURATE ---
def citeste_excel_curat(cale_fisier, coloane_implicite):
    if not os.path.exists(cale_fisier):
        return pd.DataFrame(columns=coloane_implicite)
    try:
        df = pd.read_excel(cale_fisier, skiprows=3)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        if df.empty or not all(col in df.columns for col in coloane_implicite):
            return pd.DataFrame(columns=coloane_implicite)
        return df[coloane_implicite]
    except Exception:
        return pd.DataFrame(columns=coloane_implicite)

# Generare structuri inițiale curate la prima rulare
if not os.path.exists(CENTRALIZATOR_FILE):
    salveaza_excel_stilizat(pd.DataFrame(columns=COLOANE_CENTRALIZATOR), CENTRALIZATOR_FILE, "Registru Istoric Comenzi Materiale")

if not os.path.exists(CONFIG_STOCURI_FILE):
    salveaza_excel_stilizat(pd.DataFrame(columns=COLOANE_STOCURI), CONFIG_STOCURI_FILE, "Baza de Date Stocuri Existente")


# --- PANEL ADMIN SECURIZAT (PENTRU INVENTAR/APROVIZIONARE) ---
st.sidebar.title("🔐 Panou Administrare Admin")
parola_introdusa = st.sidebar.text_input("Introdu parola Admin:", type="password")

if parola_introdusa == PAROLA_ADMIN:
    st.sidebar.success("🔑 Autentificat ca Admin!")
    st.sidebar.markdown("---")
    st.sidebar.subheader("📥 Încarcă / Re-actualizează Stocul")
    fisiere_stoc_incarcat = st.sidebar.file_uploader("Încarcă fișierul de stoc existent (XLSX sau CSV):", type=["xlsx", "csv"])
    
    if fisiere_stoc_incarcat is not None:
        try:
            if fisiere_stoc_incarcat.name.endswith('.csv'):
                df_nou = pd.read_csv(fisiere_stoc_incarcat)
            else:
                try:
                    df_nou = pd.read_excel(fisiere_stoc_incarcat, skiprows=3)
                    if 'Denumire Material' not in df_nou.columns:
                        df_nou = pd.read_excel(fisiere_stoc_incarcat)
                except:
                    df_nou = pd.read_excel(fisiere_stoc_incarcat)
            
            df_nou = df_nou.loc[:, ~df_nou.columns.str.contains('^Unnamed')]
            
            if 'Denumire Material' in df_nou.columns and 'Stoc Actual' in df_nou.columns:
                df_nou['Denumire Material'] = df_nou['Denumire Material'].astype(str).str.strip().str.upper()
                df_nou['Stoc Actual'] = pd.to_numeric(df_nou['Stoc Actual']).fillna(0).astype(int)
                if 'UM' not in df_nou.columns: df_nou['UM'] = 'buc'
                
                df_stoc_salvare = df_nou[COLOANE_STOCURI]
                
                if st.sidebar.button("💾 Încarcă noul stoc în baza de date", use_container_width=True):
                    salveaza_excel_stilizat(df_stoc_salvare, CONFIG_STOCURI_FILE, "Baza de Date Stocuri Existente")
                    st.sidebar.success("🎯 Noul stoc a fost salvat și înlocuit complet!")
                    st.rerun()
            else:
                st.sidebar.error("❌ Coloanele obligatorii în tabel trebuie să fie: 'Denumire Material' și 'Stoc Actual'")
        except Exception as e:
            st.sidebar.error(f"Eroare procesare fișier stoc: {e}")
elif parola_introdusa != "":
    st.sidebar.error("❌ Parolă incorectă!")


# --- INTERFAȚA PRINCIPALĂ (PENTRU ANGAJAȚI / UTILIZATORI) ---
st.title("🏗️ Arcaprod - Management & Control Stoc Real")

col_stanga, col_dreapta = st.columns([1, 1.3])

with col_stanga:
    st.subheader("📝 Formular Cerere Materiale")
    with st.form(key="formular_necesar", clear_on_submit=True):
        angajat = st.selectbox("Selectați Angajat:", ANGAJATI)
        tip = st.selectbox("Tip Material:", TIP_MATERIAL)
        denumire = st.text_input("Denumire Material (Ex: PANZA DEBITAT):").strip().upper()
        
        c1, c2 = st.columns(2)
        with c1:
            cantitate = st.number_input("Cantitate Solicitată:", min_value=1, step=1, format="%d")
        with c2:
            um = st.selectbox("UM:", UNITATI_MASURA)
            
        observatii = st.text_area("Observații adiționale:")
        buton_trimite = st.form_submit_button(label="Procesează și Verifică Cererea")

    if buton_trimite and denumire:
        df_stocuri = citeste_excel_curat(CONFIG_STOCURI_FILE, COLOANE_STOCURI)
        df_stocuri['Denumire Material'] = df_stocuri['Denumire Material'].astype(str).str.strip().str.upper()
        
        # Valori implicite
        stoc_initial = 0
        stoc_ramas = 0
        status_stoc = "stoc 0"
        
        if not df_stocuri.empty and denumire in df_stocuri["Denumire Material"].values:
            stoc_initial = int(df_stocuri.loc[df_stocuri["Denumire Material"] == denumire, "Stoc Actual"].values[0])
            
            if stoc_initial >= int(cantitate):
                # CAZ 1: Avem complet pe stoc
                stoc_ramas = stoc_initial - int(cantitate)
                status_stoc = "Existent in STOC"
                df_stocuri.loc[df_stocuri["Denumire Material"] == denumire, "Stoc Actual"] = stoc_ramas
            elif stoc_initial > 0:
                # CAZ 2 LOGICA NOUĂ: Avem stoc PARȚIAL (Ex: Avem 10, cere 14)
                lipsa = int(cantitate) - stoc_initial
                stoc_ramas = 0 # Le dăm pe toate cele existente, deci ramânem cu 0
                status_stoc = f"stoc 0 (S-au eliberat doar {stoc_initial} {um}, rest {lipsa} de comandat urgent!)"
                df_stocuri.loc[df_stocuri["Denumire Material"] == denumire, "Stoc Actual"] = 0
            else:
                # CAZ 3: Stocul era deja complet 0
                stoc_ramas = 0
                status_stoc = "stoc 0"
                df_stocuri.loc[df_stocuri["Denumire Material"] == denumire, "Stoc Actual"] = 0
        else:
            # Reperul nu se află în fișierul încărcat deloc
            nou_rand_stoc = pd.DataFrame([{"Denumire Material": denumire, "Stoc Actual": 0, "UM": um}])
            df_stocuri = pd.concat([df_stocuri, nou_rand_stoc], ignore_index=True)
            stoc_initial = 0
            stoc_ramas = 0
            status_stoc = "stoc 0"
            
        # Salvează fișierul de stocuri modificat
        salveaza_excel_stilizat(df_stocuri, CONFIG_STOCURI_FILE, "Baza de Date Stocuri Existente")
        
        # Generarea înregistrării în centralizator cu valorile fizice logice
        rand_nou_centralizator = {
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Angajat": angajat,
            "Tip Material": tip,
            "Denumire Material": denumire,
            "Stoc Initial": stoc_initial,
            "Cantitate Ceruta": int(cantitate),
            "Stoc Ramas": stoc_ramas,
            "UM": um,
            "Status Stoc": status_stoc,
            "Observatii": observatii
        }
        df_ex = citeste_excel_curat(CENTRALIZATOR_FILE, COLOANE_CENTRALIZATOR)
        df_final = pd.concat([df_ex, pd.DataFrame([rand_nou_centralizator])], ignore_index=True)
        salveaza_excel_stilizat(df_final, CENTRALIZATOR_FILE, "Registru Istoric Comenzi Materiale")
        
        if status_stoc == "Existent in STOC":
            st.success(f"✅ Material aprobat complet. Status: {status_stoc}. (Stoc rămas în magazin: {stoc_ramas} {um})")
        else:
            st.error(f"🚨 {status_stoc}")

with col_dreapta:
    df_hist = citeste_excel_curat(CENTRALIZATOR_FILE, COLOANE_CENTRALIZATOR)
    df_stocuri = citeste_excel_curat(CONFIG_STOCURI_FILE, COLOANE_STOCURI)
    
    st.subheader("📧 Expediere Rapoarte")
    if st.button("🚀 Trimite fișierele Excel către sageata14@gmail.com", use_container_width=True):
        with st.spinner("Se trimit fișierele..."):
            if trimite_email_cu_atansamente([CENTRALIZATOR_FILE, CONFIG_STOCURI_FILE]):
                st.success(f"📩 Rapoartele Excel au fost livrate cu succes la {EMAIL_DESTINATAR}!")

    st.markdown("---")

    st.subheader("📊 Inventar Depozit în Timp Real")
    if not df_stocuri.empty:
        df_stocuri["Stoc Actual"] = pd.to_numeric(df_stocuri["Stoc Actual"]).fillna(0).astype(int)
        df_afisare_stoc = df_stocuri.copy()
        df_afisare_stoc["Status Material"] = df_afisare_stoc["Stoc Actual"].apply(lambda x: "✅ Disponibil în Depozit" if x > 0 else "🚨 STOC 0")
        st.dataframe(df_afisare_stoc, use_container_width=True, hide_index=True)
    else:
        st.info("Baza de date este goală. Te rog accesează Panoul de Admin din stânga pentru a încărca stocul inițial.")

    st.markdown("---")

    st.subheader("📋 Istoric Cereri și Mișcări de Stoc")
    if not df_hist.empty:
        if datetime.now().hour >= 15:
            st.warning("🔒 După ora 15:00, registrul de modificări zilnice este securizat.")
        st.dataframe(df_hist, use_container_width=True, hide_index=True)
    else:
        st.info("Nicio cerere operată în sesiunea curentă.")
